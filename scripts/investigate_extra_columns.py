"""
investigate_extra_columns.py
-----------------------------
Deep investigation of ALL extra columns (beyond core 7) in individual sprint tabs.
Goal: Catalog every additional data field, its fill rate, sample values, data type,
and cross-tab consistency. These columns are NOT manually enriched — they come from
the DB and are potential features for NPS modeling.
"""

import sys
import io
import os
from datetime import datetime
from collections import Counter, OrderedDict

# UTF-8 output for Hindi text
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

# ---------- Config ----------
EXCEL_PATH = r"C:\Users\nikhi\Downloads\NPS Verma Parivar.xlsx"
OUTPUT_DIR = r"C:\Users\nikhi\wiom-nps-analysis\output"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "extra_columns_investigation.txt")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Tabs to investigate (representative sample across time periods)
TARGET_TABS = [
    'Sprint 7 Oct25',
    'Sprint 14 Jan26',
    'Sprint RSP1 Feb26',
    'Sprint RSP2 Feb26',
    'Sprint RSP3 Feb26',
]

CORE_COLUMNS = [
    'USER_ID', 'PROFILE_ALL_IDENTITIES', 'USER_RATING',
    'NPS_CLASSIFICATION', 'COMMENT', 'TIMESTAMP', 'Mobile'
]

HEADER_ROW = 14  # Known header row for these sprint tabs

lines: list[str] = []


def log(msg: str = ""):
    lines.append(msg)
    print(msg)


def log_sep(char="-", width=100):
    log(char * width)


def infer_dtype(values):
    """Infer the dominant data type from a list of non-null values."""
    if not values:
        return "empty"

    type_counts = Counter()
    for v in values:
        if v is None:
            continue
        if isinstance(v, (int, float)):
            # Check if it looks like an integer stored as float
            if isinstance(v, float) and v == int(v):
                type_counts['integer'] += 1
            else:
                type_counts['numeric'] += 1
        elif isinstance(v, datetime):
            type_counts['datetime'] += 1
        elif isinstance(v, bool):
            type_counts['boolean'] += 1
        elif isinstance(v, str):
            s = str(v).strip()
            if not s:
                continue
            # Check for date-like strings
            if any(sep in s for sep in ['/', '-']) and any(c.isdigit() for c in s):
                try:
                    # Quick check — if it has T and : likely a timestamp
                    if 'T' in s and ':' in s:
                        type_counts['datetime_string'] += 1
                        continue
                except:
                    pass
            # Check for numeric strings
            try:
                float(s)
                type_counts['numeric_string'] += 1
                continue
            except ValueError:
                pass
            # Check for yes/no/true/false
            if s.lower() in ('yes', 'no', 'true', 'false', 'y', 'n', '1', '0',
                              'haan', 'nahi', 'na'):
                type_counts['boolean_like'] += 1
            else:
                type_counts['text'] += 1
        else:
            type_counts[type(v).__name__] += 1

    if not type_counts:
        return "empty"

    # Return dominant type, with detail
    dominant = type_counts.most_common(1)[0][0]
    total = sum(type_counts.values())

    if len(type_counts) == 1:
        return dominant
    else:
        parts = [f"{t}({c})" for t, c in type_counts.most_common()]
        return f"{dominant} [mixed: {', '.join(parts)}]"


def safe_str(v, max_len=80):
    """Convert value to display string, truncating if needed."""
    if v is None:
        return "<null>"
    s = str(v).strip()
    if not s:
        return "<empty>"
    # Replace newlines for display
    s = s.replace('\n', ' | ').replace('\r', '')
    if len(s) > max_len:
        s = s[:max_len - 3] + "..."
    return s


def get_unique_values_summary(values, max_unique=20):
    """Get a summary of unique values for categorical columns."""
    non_null = [v for v in values if v is not None and str(v).strip()]
    if not non_null:
        return "No non-null values"

    counter = Counter(str(v).strip() for v in non_null)
    total = len(non_null)
    unique_count = len(counter)

    summary_parts = []
    for val, cnt in counter.most_common(max_unique):
        pct = cnt / total * 100
        display_val = val[:50] + "..." if len(val) > 50 else val
        summary_parts.append(f"  '{display_val}' = {cnt} ({pct:.1f}%)")

    header = f"  Unique values: {unique_count} | Total non-null: {total}"
    if unique_count > max_unique:
        summary_parts.append(f"  ... and {unique_count - max_unique} more unique values")

    return header + "\n" + "\n".join(summary_parts)


# ---------- Main ----------

def main():
    log("=" * 100)
    log("EXTRA COLUMNS DEEP INVESTIGATION")
    log(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log(f"Source: {EXCEL_PATH}")
    log(f"Target tabs: {TARGET_TABS}")
    log("=" * 100)

    log("\nLoading workbook (this takes a minute)...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False, data_only=True)

    # Store results per tab
    tab_results = OrderedDict()  # tab_name -> {col_name: {fill_rate, samples, dtype, values}}

    # ============================================================
    # SECTION 1: Analyze each target tab
    # ============================================================
    for tab_name in TARGET_TABS:
        log_sep("=")
        log(f"\n>>> TAB: '{tab_name}' <<<")
        log_sep("=")

        if tab_name not in wb.sheetnames:
            log(f"  WARNING: Tab '{tab_name}' not found in workbook!")
            continue

        ws = wb[tab_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        log(f"  Dimensions: {max_row} rows x {max_col} cols")

        # Read headers from row 14
        headers = []
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=HEADER_ROW, column=col_idx).value
            if val is not None:
                headers.append(str(val).strip())
            else:
                # Check if there are more non-null headers after this
                # (sometimes there are gaps)
                headers.append(None)

        # Trim trailing None headers
        while headers and headers[-1] is None:
            headers.pop()

        log(f"  Total columns (including core): {len(headers)}")
        log(f"  Core columns (1-7): {headers[:7]}")

        # Read all data rows
        data_rows = []
        for row_idx in range(HEADER_ROW + 1, max_row + 1):
            row_vals = []
            all_none = True
            for col_idx in range(1, len(headers) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_vals.append(val)
                if val is not None:
                    all_none = False
            if not all_none:
                data_rows.append(row_vals)

        total_rows = len(data_rows)
        log(f"  Data rows: {total_rows}")

        # Identify extra columns (beyond core 7)
        extra_cols = []
        for i, h in enumerate(headers):
            if i >= 7 and h is not None:  # Beyond core 7
                extra_cols.append((i, h))

        log(f"  Extra columns count: {len(extra_cols)}")
        log(f"\n  ALL COLUMN NAMES (including core):")
        for i, h in enumerate(headers):
            marker = " [CORE]" if i < 7 else " [EXTRA]"
            display_h = h if h else f"<unnamed_col_{i+1}>"
            log(f"    Col {i+1:2d}: {display_h}{marker}")

        # Analyze each extra column in detail
        tab_col_data = OrderedDict()

        log(f"\n  {'='*90}")
        log(f"  DETAILED EXTRA COLUMN ANALYSIS")
        log(f"  {'='*90}")

        for col_idx, col_name in extra_cols:
            # Extract all values for this column
            col_values = [row[col_idx] if col_idx < len(row) else None for row in data_rows]

            # Non-null values
            non_null = [v for v in col_values if v is not None and str(v).strip() not in ('', 'None')]
            fill_count = len(non_null)
            fill_rate = fill_count / total_rows * 100 if total_rows > 0 else 0

            # First 5 non-null sample values
            samples = []
            for v in non_null[:5]:
                samples.append(safe_str(v))

            # Data type inference
            dtype = infer_dtype(non_null[:200])  # Sample first 200 for performance

            # Store for cross-tab comparison
            tab_col_data[col_name] = {
                'fill_rate': fill_rate,
                'fill_count': fill_count,
                'total_rows': total_rows,
                'samples': samples,
                'dtype': dtype,
                'all_values': non_null,
            }

            log(f"\n  --- Column: '{col_name}' (position {col_idx + 1}) ---")
            log(f"      Fill rate: {fill_rate:.1f}% ({fill_count}/{total_rows})")
            log(f"      Data type: {dtype}")
            log(f"      Sample values (first 5 non-null):")
            for s in samples:
                log(f"        - {s}")

            # For low-cardinality columns, show value distribution
            unique_non_null = set(str(v).strip() for v in non_null if str(v).strip())
            if len(unique_non_null) <= 30:
                log(f"      Value distribution:")
                log(f"  {get_unique_values_summary(non_null)}")

        tab_results[tab_name] = tab_col_data

    # ============================================================
    # SECTION 2: Specific column deep dives
    # ============================================================
    log("\n" + "=" * 100)
    log(">>> SPECIFIC COLUMN DEEP DIVES <<<")
    log("=" * 100)

    # List of columns to specifically investigate
    specific_columns = {
        'pay g?': 'Payment gateway / payment method flag',
        'cash/ online': 'Cash vs online payment mode',
        'Tenure': 'Customer tenure / duration of service',
        'City': 'Customer city / geography',
        'Optical Power': 'Fiber optical power readings (signal strength)',
        'recharge done': 'Whether recharge was completed',
        '# of tickets in last 3 months': 'Support ticket count',
        '# of tx in last 3 months': 'Support ticket count (alternate name)',
        'Device Type': 'Router/device type',
        'Device type': 'Router/device type (alternate case)',
        'Device': 'Device model',
        'Zone': 'Geographic zone',
        'ALT mobile': 'Alternate mobile number',
        'Devices on 2.4g': 'Count of 2.4GHz devices',
        'Devices on 5g': 'Count of 5GHz devices',
        'Payment_issue\nTickets Flag': 'Payment issue ticket flag',
        'Number of Devices on Wiom Net': 'Total devices connected',
    }

    for target_col, description in specific_columns.items():
        log_sep("-")
        log(f"\n  COLUMN: '{target_col}' ({description})")

        found_in = []
        for tab_name, col_data in tab_results.items():
            # Try exact match first, then case-insensitive, then partial
            matched_key = None
            for key in col_data.keys():
                if key == target_col:
                    matched_key = key
                    break
            if not matched_key:
                for key in col_data.keys():
                    if key and target_col and key.lower().strip() == target_col.lower().strip():
                        matched_key = key
                        break
            if not matched_key:
                # Partial match
                for key in col_data.keys():
                    if key and target_col:
                        tl = target_col.lower().replace('\n', ' ').strip()
                        kl = key.lower().replace('\n', ' ').strip()
                        if tl in kl or kl in tl:
                            matched_key = key
                            break

            if matched_key:
                info = col_data[matched_key]
                found_in.append((tab_name, matched_key, info))

        if not found_in:
            log(f"    NOT FOUND in any of the target tabs")
        else:
            log(f"    Found in {len(found_in)}/{len(tab_results)} tabs:")
            for tab_name, actual_key, info in found_in:
                log(f"\n    Tab: '{tab_name}' (as '{actual_key}')")
                log(f"      Fill rate: {info['fill_rate']:.1f}%")
                log(f"      Data type: {info['dtype']}")
                log(f"      Samples: {info['samples'][:5]}")

                # Show full value distribution for key columns
                vals = info['all_values']
                if vals:
                    unique_vals = set(str(v).strip() for v in vals if str(v).strip())
                    if len(unique_vals) <= 25:
                        counter = Counter(str(v).strip() for v in vals if str(v).strip())
                        log(f"      Full value distribution ({len(unique_vals)} unique):")
                        for val, cnt in counter.most_common():
                            pct = cnt / len(vals) * 100
                            log(f"        '{val}' = {cnt} ({pct:.1f}%)")
                    else:
                        log(f"      High cardinality: {len(unique_vals)} unique values")
                        # Show numeric stats if numeric
                        try:
                            numeric_vals = [float(str(v).strip()) for v in vals
                                          if str(v).strip() and str(v).strip() not in ('#N/A', '#VALUE!', '#REF!', 'NA', 'N/A')]
                            if numeric_vals:
                                log(f"      Numeric stats: min={min(numeric_vals):.2f}, max={max(numeric_vals):.2f}, "
                                    f"mean={sum(numeric_vals)/len(numeric_vals):.2f}, count={len(numeric_vals)}")
                        except (ValueError, TypeError):
                            pass

    # ============================================================
    # SECTION 3: Cross-tab consistency analysis
    # ============================================================
    log("\n" + "=" * 100)
    log(">>> CROSS-TAB CONSISTENCY ANALYSIS <<<")
    log("=" * 100)

    # Collect all extra column names across all tabs (normalized)
    all_extra_cols = OrderedDict()  # normalized_name -> {tabs: [tab_names], original_names: [names]}

    for tab_name, col_data in tab_results.items():
        for col_name in col_data.keys():
            # Normalize: lowercase, strip, replace newlines
            norm = col_name.lower().strip().replace('\n', ' ').replace('\r', '')
            if norm not in all_extra_cols:
                all_extra_cols[norm] = {
                    'tabs': [],
                    'original_names': [],
                    'fill_rates': [],
                }
            all_extra_cols[norm]['tabs'].append(tab_name)
            all_extra_cols[norm]['original_names'].append(col_name)
            all_extra_cols[norm]['fill_rates'].append(col_data[col_name]['fill_rate'])

    num_tabs = len(tab_results)

    # Columns present in ALL tabs
    universal_cols = {k: v for k, v in all_extra_cols.items() if len(v['tabs']) == num_tabs}
    # Columns present in SOME tabs
    partial_cols = {k: v for k, v in all_extra_cols.items() if 1 < len(v['tabs']) < num_tabs}
    # Columns present in only ONE tab
    unique_cols = {k: v for k, v in all_extra_cols.items() if len(v['tabs']) == 1}

    log(f"\n  Total unique extra columns (normalized): {len(all_extra_cols)}")
    log(f"  Present in ALL {num_tabs} tabs: {len(universal_cols)}")
    log(f"  Present in SOME tabs (2-{num_tabs-1}): {len(partial_cols)}")
    log(f"  Present in only ONE tab: {len(unique_cols)}")

    log(f"\n  COLUMNS PRESENT IN ALL {num_tabs} TABS (universal):")
    log(f"  {'Column Name':<45} {'Avg Fill%':>10} {'Name Variants'}")
    log(f"  {'-'*45} {'-'*10} {'-'*40}")
    for norm_name, info in sorted(universal_cols.items()):
        avg_fill = sum(info['fill_rates']) / len(info['fill_rates'])
        # Check for name variations
        unique_names = list(set(info['original_names']))
        name_display = unique_names[0] if len(unique_names) == 1 else f"{unique_names}"
        log(f"  {norm_name:<45} {avg_fill:>9.1f}% {name_display}")

    if partial_cols:
        log(f"\n  COLUMNS PRESENT IN SOME TABS (partial):")
        log(f"  {'Column Name':<45} {'In Tabs':>8} {'Avg Fill%':>10} {'Present In'}")
        log(f"  {'-'*45} {'-'*8} {'-'*10} {'-'*40}")
        for norm_name, info in sorted(partial_cols.items(), key=lambda x: -len(x[1]['tabs'])):
            avg_fill = sum(info['fill_rates']) / len(info['fill_rates'])
            short_tabs = [t.replace('Sprint ', 'S').replace(' Feb26', '').replace(' Jan26', '').replace(' Oct25', '')
                         for t in info['tabs']]
            log(f"  {norm_name:<45} {len(info['tabs']):>5}/{num_tabs} {avg_fill:>9.1f}% {short_tabs}")

    if unique_cols:
        log(f"\n  COLUMNS PRESENT IN ONLY ONE TAB:")
        log(f"  {'Column Name':<45} {'Tab':>25} {'Fill%':>8}")
        log(f"  {'-'*45} {'-'*25} {'-'*8}")
        for norm_name, info in sorted(unique_cols.items()):
            tab = info['tabs'][0].replace('Sprint ', 'S')
            fill = info['fill_rates'][0]
            log(f"  {norm_name:<45} {tab:>25} {fill:>7.1f}%")

    # ============================================================
    # SECTION 4: Check Consolidated tab for equivalents
    # ============================================================
    log("\n" + "=" * 100)
    log(">>> CONSOLIDATED TAB EQUIVALENTS CHECK <<<")
    log("=" * 100)

    consol_sheet = 'Consolidated'
    if consol_sheet in wb.sheetnames:
        ws_c = wb[consol_sheet]
        # Consolidated has header at row 1
        consol_headers = []
        max_col_c = ws_c.max_column or 0
        for col_idx in range(1, max_col_c + 1):
            val = ws_c.cell(row=1, column=col_idx).value
            if val is not None:
                consol_headers.append(str(val).strip())
            else:
                consol_headers.append(None)

        # Trim trailing Nones
        while consol_headers and consol_headers[-1] is None:
            consol_headers.pop()

        log(f"\n  Consolidated tab has {len(consol_headers)} columns:")
        for i, h in enumerate(consol_headers):
            log(f"    Col {i+1:2d}: {h}")

        # Now map sprint extra columns to Consolidated equivalents
        log(f"\n  MAPPING: Sprint Extra Columns -> Consolidated Equivalents")
        log(f"  {'Sprint Column':<45} {'Consolidated Equivalent':<40} {'Match Type'}")
        log(f"  {'-'*45} {'-'*40} {'-'*15}")

        consol_norms = {str(h).lower().strip().replace('\n', ' '): h
                       for h in consol_headers if h}

        for norm_name, info in sorted(all_extra_cols.items()):
            original = info['original_names'][0]

            # Try to find equivalent in consolidated
            matched = None
            match_type = "NONE"

            # Exact match
            if norm_name in consol_norms:
                matched = consol_norms[norm_name]
                match_type = "exact"
            else:
                # Partial/fuzzy match
                for cn, cv in consol_norms.items():
                    # Check containment both ways
                    if norm_name in cn or cn in norm_name:
                        matched = cv
                        match_type = "partial"
                        break
                    # Check key words
                    norm_words = set(norm_name.split())
                    cn_words = set(cn.split())
                    overlap = norm_words & cn_words
                    if len(overlap) >= 2 or (len(overlap) == 1 and
                        list(overlap)[0] not in ('of', 'in', 'on', 'the', 'a', 'is', 'and')):
                        if list(overlap)[0] in ('tenure', 'city', 'zone', 'device', 'recharge',
                                                 'ticket', 'payment', 'optical', 'power', 'churn'):
                            matched = cv
                            match_type = "keyword"
                            break

            equiv_display = matched if matched else "---"
            tabs_count = len(info['tabs'])
            log(f"  {original:<45} {equiv_display:<40} {match_type} (in {tabs_count}/{num_tabs} tabs)")

    # ============================================================
    # SECTION 5: Modeling-useful column summary
    # ============================================================
    log("\n" + "=" * 100)
    log(">>> MODELING-USEFUL COLUMNS SUMMARY <<<")
    log("=" * 100)

    log("""
  These extra columns from sprint tabs are potentially useful as features for NPS
  prediction/analysis. They appear to come from the database (not manually enriched).

  CATEGORIES OF USEFUL FEATURES:
  """)

    # Categorize columns
    categories = {
        'CUSTOMER PROFILE': [],
        'PAYMENT & BILLING': [],
        'NETWORK & TECHNICAL': [],
        'GEOGRAPHY': [],
        'SUPPORT / TICKETS': [],
        'DEVICE INFO': [],
        'ENGAGEMENT / BEHAVIOR': [],
        'CALLING / OUTREACH': [],
        'UNKNOWN / OTHER': [],
    }

    for norm_name, info in all_extra_cols.items():
        original = info['original_names'][0]
        avg_fill = sum(info['fill_rates']) / len(info['fill_rates'])
        tabs_count = len(info['tabs'])

        entry = f"{original} (fill: {avg_fill:.0f}%, in {tabs_count}/{num_tabs} tabs)"

        nl = norm_name.lower()
        if any(w in nl for w in ['tenure', 'active', 'churn', 'install']):
            categories['CUSTOMER PROFILE'].append(entry)
        elif any(w in nl for w in ['pay', 'cash', 'online', 'recharge', 'billing']):
            categories['PAYMENT & BILLING'].append(entry)
        elif any(w in nl for w in ['optical', 'power', 'signal', 'bandwidth', '2.4g', '5g',
                                     'frequency', 'speed', 'latency', 'uptime', 'rssi']):
            categories['NETWORK & TECHNICAL'].append(entry)
        elif any(w in nl for w in ['city', 'zone', 'area', 'region', 'location', 'geo']):
            categories['GEOGRAPHY'].append(entry)
        elif any(w in nl for w in ['ticket', 'complaint', 'issue', 'resolution']):
            categories['SUPPORT / TICKETS'].append(entry)
        elif any(w in nl for w in ['device', 'router', 'model', 'hardware']):
            categories['DEVICE INFO'].append(entry)
        elif any(w in nl for w in ['call', 'contact', 'attempt', 'reach', 'answer',
                                     'outbound', 'det callin', 'connected']):
            categories['CALLING / OUTREACH'].append(entry)
        elif any(w in nl for w in ['mobile', 'alt', 'number']):
            categories['CUSTOMER PROFILE'].append(entry)
        elif any(w in nl for w in ['tag', 'category', 'reason', 'nps', 'comment',
                                     'classification', 'score']):
            categories['ENGAGEMENT / BEHAVIOR'].append(entry)
        else:
            categories['UNKNOWN / OTHER'].append(entry)

    for cat_name, entries in categories.items():
        if entries:
            log(f"  {cat_name}:")
            for e in entries:
                log(f"    - {e}")
            log("")

    # ============================================================
    # SECTION 6: Also check rows 1-13 (summary area) structure
    # ============================================================
    log("\n" + "=" * 100)
    log(">>> ROWS 1-13 SUMMARY AREA STRUCTURE <<<")
    log("=" * 100)
    log("  (Rows 1-13 contain pre-computed analysis tables above the data)")

    # Just check Sprint RSP3 as the most recent/complete
    check_tab = 'Sprint RSP3 Feb26'
    if check_tab in wb.sheetnames:
        ws_s = wb[check_tab]
        log(f"\n  Sample from '{check_tab}' rows 1-13:")
        for row_idx in range(1, 14):
            row_vals = []
            for col_idx in range(1, min(10, (ws_s.max_column or 10) + 1)):
                val = ws_s.cell(row=row_idx, column=col_idx).value
                row_vals.append(safe_str(val, 30) if val is not None else "")
            # Only show if row has content
            if any(v for v in row_vals):
                log(f"    Row {row_idx:2d}: {row_vals}")

    # ============================================================
    # Final summary
    # ============================================================
    log("\n" + "=" * 100)
    log(">>> FINAL SUMMARY & KEY FINDINGS <<<")
    log("=" * 100)

    log(f"""
  TABS ANALYZED: {len(tab_results)}
  TOTAL EXTRA COLUMNS FOUND (unique, normalized): {len(all_extra_cols)}

  KEY FINDINGS:
  1. Universal columns (in all tabs): {len(universal_cols)}
  2. Partial columns (in some tabs):  {len(partial_cols)}
  3. Tab-specific columns:            {len(unique_cols)}

  HIGH-VALUE FEATURES FOR MODELING (from DB, not manual):
  - Tenure: customer duration (categorical: 1-2 months, 3-6 months, 6+ months)
  - City: geography for regional analysis
  - Zone: geographic zone grouping
  - Device Type: router model (GX, SY, TP, etc.)
  - Optical Power: fiber signal strength (continuous)
  - Recharge Done: payment behavior flag
  - Cash/Online: payment mode
  - # Tickets: support burden measure
  - Devices on 2.4g / 5g: network usage complexity
  - Payment Issue Tickets: billing-specific support

  CRITICAL NOTE: The user confirms these columns are NOT manually enriched.
  They come from the database, making them reliable features for modeling.
  They can be joined to any NPS response via the Mobile/phone number.
""")

    wb.close()

    # Write report
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"\n[Report saved to: {OUTPUT_FILE}]")


if __name__ == "__main__":
    main()
