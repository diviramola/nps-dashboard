"""
investigate_sprint_tabs.py
--------------------------
Read-only investigation of ALL sprint/NPS tabs in the Excel workbook.
Goal: Find how much data exists beyond the "Consolidated" tab.

Key insight from first run: Sprint tabs have header at row 14 with columns:
  USER_ID | PROFILE_ALL_IDENTITIES | USER_RATING | NPS_CLASSIFICATION | COMMENT | TIMESTAMP | Mobile
  plus additional analysis columns (device, tenure, city, etc.) from col 8 onward.
The phone column is called "Mobile", not "Phone Number".
Rows 1-13 are analysis/summary rows (NPS breakdown, device split, city split).
"""

import sys
import io
import os
from datetime import datetime

# UTF-8 output for Hindi text
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from collections import OrderedDict, Counter

# ---------- Config ----------
EXCEL_PATH = r"C:\Users\nikhi\Downloads\NPS Verma Parivar.xlsx"
OUTPUT_DIR = r"C:\Users\nikhi\wiom-nps-analysis\output"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "sprint_tabs_investigation.txt")
os.makedirs(OUTPUT_DIR, exist_ok=True)

lines: list[str] = []

def log(msg: str = ""):
    lines.append(msg)
    print(msg)

def log_sep(char="-", width=90):
    log(char * width)

# ---------- Helpers ----------

def safe_max_col(ws):
    return ws.max_column or 20

def safe_max_row(ws):
    return ws.max_row or 0


def find_header_row(ws, max_scan=60):
    """
    Scan for header row. Look for:
    1. A row containing 'Phone Number' or 'Phone' (Consolidated style)
    2. A row containing 'Mobile' as a header (Sprint style)
    3. A row containing 'USER_ID' (Sprint raw data header)
    Returns (header_row_index_1based, list_of_header_values).
    """
    mc = safe_max_col(ws)
    mr = safe_max_row(ws)
    scan_limit = min(max_scan, mr)

    for row_idx in range(1, scan_limit + 1):
        row_vals = []
        for col_idx in range(1, mc + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            row_vals.append(val)

        # Check for known header patterns
        str_vals = [str(v).strip().lower() if v is not None else '' for v in row_vals]

        # Pattern 1: Contains 'phone number' (Consolidated)
        if any('phone number' in s for s in str_vals):
            return row_idx, row_vals

        # Pattern 2: 'USER_ID' in first column and 'Mobile' somewhere (Sprint tabs)
        if str_vals and str_vals[0] == 'user_id':
            return row_idx, row_vals

        # Pattern 3: Has 'mobile' as a standalone header value
        if 'mobile' in str_vals:
            # Make sure it looks like a header row (not data)
            # Header rows typically have text in most cells
            non_empty = sum(1 for s in str_vals if s)
            if non_empty >= 3:
                return row_idx, row_vals

    return None, None


def extract_data_rows(ws, header_row_idx, headers):
    """Read all non-empty data rows starting from header_row_idx + 1."""
    data = []
    clean_headers = []
    for h in headers:
        if h is None:
            clean_headers.append(f"_col_{len(clean_headers)+1}")
        else:
            clean_headers.append(str(h).strip())

    mr = safe_max_row(ws)
    for row_idx in range(header_row_idx + 1, mr + 1):
        row_vals = []
        all_none = True
        for col_idx in range(1, len(clean_headers) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            row_vals.append(val)
            if val is not None:
                all_none = False
        if all_none:
            continue
        row_dict = dict(zip(clean_headers, row_vals))
        data.append(row_dict)
    return data, clean_headers


def find_phone_column(headers):
    """Find the column name for phone/mobile numbers."""
    for h in headers:
        if h and isinstance(h, str):
            hl = h.lower().strip()
            if 'phone number' in hl:
                return h
            if hl == 'mobile':
                return h
            if hl == 'phone':
                return h
    return None


def extract_phones(data, phone_col):
    """Extract set of unique phone numbers (as strings, normalized)."""
    phones = set()
    if not phone_col:
        return phones
    for row in data:
        val = row.get(phone_col)
        if val is not None:
            s = str(val).strip()
            # Clean up: remove .0 suffix from float conversion
            if s.endswith('.0'):
                s = s[:-2]
            # Skip empty, none, error values
            if s and s.lower() not in ('none', '', '#value!', '#ref!', '#n/a'):
                # Should look like a phone number (mostly digits, 10+ chars)
                digits_only = ''.join(c for c in s if c.isdigit())
                if len(digits_only) >= 10:
                    # Normalize to last 10 digits for comparison
                    phones.add(digits_only[-10:])
    return phones


def extract_nps_score_column(headers):
    """Find NPS score column."""
    for h in headers:
        if h and isinstance(h, str):
            hl = h.lower().strip()
            if hl == 'nps':
                return h
            if hl == 'user_rating':
                return h
            if 'nps' in hl and 'reason' not in hl and 'group' not in hl and 'class' not in hl:
                return h
            if hl in ('score', 'rating'):
                return h
    return None


def extract_nps_classification_column(headers):
    """Find NPS classification column (Promoter/Passive/Detractor)."""
    for h in headers:
        if h and isinstance(h, str):
            hl = h.lower().strip()
            if hl == 'nps_classification':
                return h
            if hl == 'nps group':
                return h
    return None


# ---------- Main ----------

def main():
    log("=" * 90)
    log("SPRINT TABS INVESTIGATION REPORT (v2 - fixed Mobile column detection)")
    log(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log(f"Source: {EXCEL_PATH}")
    log("=" * 90)

    log("\nLoading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False, data_only=True)
    all_sheets = wb.sheetnames

    log(f"\nTotal sheets in workbook: {len(all_sheets)}")
    log("\nAll sheet names:")
    for i, name in enumerate(all_sheets, 1):
        log(f"  {i:3d}. '{name}'")

    # ---------- Classify sheets ----------
    sprint_data_sheets = []  # Sprint NPS tabs with survey data
    consolidated_sheet = None
    other_sheets = []

    for name in all_sheets:
        nl = name.lower().strip()
        if 'consolidated' in nl:
            consolidated_sheet = name
        elif nl.startswith('sprint') and 'dump' not in nl:
            sprint_data_sheets.append(name)
        else:
            other_sheets.append(name)

    log_sep("=")
    log("\nSHEET CLASSIFICATION:")
    log(f"  Consolidated tab : '{consolidated_sheet}'")
    log(f"  Sprint DATA tabs : {len(sprint_data_sheets)}")
    log(f"  Other tabs       : {len(other_sheets)}")

    log("\n  Sprint DATA sheets (primary analysis targets):")
    for name in sprint_data_sheets:
        log(f"    - '{name}'")

    # ---------- Read Consolidated tab ----------
    log_sep("=")
    log("\n>>> CONSOLIDATED TAB ANALYSIS <<<")
    log_sep()

    consol_phones = set()
    consol_data = []
    consol_headers = []
    consol_phone_col = None
    if consolidated_sheet:
        ws = wb[consolidated_sheet]
        hdr_row, hdr_vals = find_header_row(ws)
        if hdr_row:
            consol_data, consol_headers = extract_data_rows(ws, hdr_row, hdr_vals)
            consol_phone_col = find_phone_column(consol_headers)
            consol_phones = extract_phones(consol_data, consol_phone_col)
            log(f"  Sheet: '{consolidated_sheet}'")
            log(f"  Header row: {hdr_row}")
            log(f"  Key columns: {consol_headers[:15]}...")
            log(f"  Total columns: {len(consol_headers)}")
            log(f"  Data rows: {len(consol_data)}")
            log(f"  Phone column: '{consol_phone_col}'")
            log(f"  Unique phones (normalized 10-digit): {len(consol_phones)}")

            # Show NPS distribution from Consolidated
            nps_col = extract_nps_score_column(consol_headers)
            nps_class_col = extract_nps_classification_column(consol_headers)
            if nps_class_col:
                class_counts = Counter(str(row.get(nps_class_col, '')).strip() for row in consol_data)
                log(f"  NPS Classification distribution:")
                for cls, cnt in class_counts.most_common():
                    if cls:
                        log(f"    {cls}: {cnt}")

            # Check sprint range in Consolidated
            sprint_id_col = None
            for h in consol_headers:
                if h and 'sprint id' in str(h).lower():
                    sprint_id_col = h
                    break
            if sprint_id_col:
                sprint_ids = set(str(row.get(sprint_id_col, '')).strip() for row in consol_data if row.get(sprint_id_col))
                log(f"  Sprint IDs in Consolidated: {sorted(sprint_ids)}")
        else:
            log(f"  WARNING: Could not find header row in '{consolidated_sheet}'")
    else:
        log("  WARNING: No Consolidated sheet found!")

    # ---------- Analyze each sprint DATA tab ----------
    log_sep("=")
    log("\n>>> INDIVIDUAL SPRINT DATA TAB ANALYSIS <<<")
    log_sep("=")

    sprint_tab_info = OrderedDict()
    all_sprint_phones = set()
    total_sprint_rows = 0
    missing_from_consolidated = set()

    rsp_tabs = []
    jan26_tabs = []

    for sheet_name in sprint_data_sheets:
        log_sep()
        log(f"\n  SHEET: '{sheet_name}'")
        log(f"  {'~' * (len(sheet_name) + 10)}")

        nl = sheet_name.lower()
        if 'rsp' in nl:
            rsp_tabs.append(sheet_name)
        if '14' in nl and 'jan' in nl:
            jan26_tabs.append(sheet_name)

        ws = wb[sheet_name]
        mr = safe_max_row(ws)
        mc = safe_max_col(ws)
        log(f"  Raw dimensions: {mr} rows x {mc} cols")

        hdr_row, hdr_vals = find_header_row(ws, max_scan=60)

        info = {
            'header_row': hdr_row,
            'headers': [],
            'core_headers': [],  # First 7 standard columns
            'extra_headers': [],  # Analysis columns beyond 7
            'data_rows': 0,
            'unique_phones': 0,
            'phones': set(),
            'phones_not_in_consolidated': set(),
            'phone_col': None,
            'nps_col': None,
            'nps_class_col': None,
            'nps_distribution': {},
        }

        if hdr_row:
            data, headers = extract_data_rows(ws, hdr_row, hdr_vals)
            phone_col = find_phone_column(headers)
            nps_col = extract_nps_score_column(headers)
            nps_class_col = extract_nps_classification_column(headers)
            phones = extract_phones(data, phone_col)
            phones_missing = phones - consol_phones

            info['headers'] = headers
            info['core_headers'] = headers[:7] if len(headers) >= 7 else headers
            info['extra_headers'] = headers[7:] if len(headers) > 7 else []
            info['data_rows'] = len(data)
            info['unique_phones'] = len(phones)
            info['phones'] = phones
            info['phones_not_in_consolidated'] = phones_missing
            info['phone_col'] = phone_col
            info['nps_col'] = nps_col
            info['nps_class_col'] = nps_class_col

            if nps_class_col:
                info['nps_distribution'] = dict(Counter(
                    str(row.get(nps_class_col, '')).strip() for row in data if row.get(nps_class_col)
                ))

            all_sprint_phones.update(phones)
            total_sprint_rows += len(data)
            missing_from_consolidated.update(phones_missing)

            log(f"  Header row: {hdr_row}")
            log(f"  Core columns (1-7): {info['core_headers']}")
            log(f"  Extra analysis columns: {len(info['extra_headers'])}")
            if info['extra_headers']:
                log(f"    Extra cols: {info['extra_headers'][:10]}{'...' if len(info['extra_headers']) > 10 else ''}")
            log(f"  Data rows: {len(data)}")
            log(f"  Phone column: '{phone_col}'")
            log(f"  NPS score column: '{nps_col}'")
            log(f"  NPS classification column: '{nps_class_col}'")
            log(f"  Unique phones: {len(phones)}")
            log(f"  Phones NOT in Consolidated: {len(phones_missing)}")

            if info['nps_distribution']:
                log(f"  NPS distribution:")
                for cls, cnt in sorted(info['nps_distribution'].items(), key=lambda x: -x[1]):
                    log(f"    {cls}: {cnt}")

            # Show sample data
            if data[:2]:
                log(f"\n  Sample data (first 2 rows, core cols only):")
                for i, row in enumerate(data[:2], 1):
                    log(f"    Row {i}:")
                    for k in info['core_headers']:
                        val_str = str(row.get(k, '')) if row.get(k) is not None else "<empty>"
                        if len(val_str) > 60:
                            val_str = val_str[:57] + "..."
                        log(f"      {k}: {val_str}")
        else:
            log(f"\n  WARNING: No header row found in first 60 rows.")
            log(f"  Showing first 15 rows to understand structure:")
            for row_idx in range(1, min(16, mr + 1)):
                row_vals = []
                for col_idx in range(1, min(8, mc + 1)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    row_vals.append(val)
                display_vals = [str(v)[:35] if v else "" for v in row_vals]
                log(f"    Row {row_idx:3d}: {display_vals}")

        sprint_tab_info[sheet_name] = info

    # ---------- Special focus: RSP tabs ----------
    log_sep("=")
    log("\n>>> SPECIAL FOCUS: 'Sprint RSP...' TABS <<<")
    log_sep()

    if rsp_tabs:
        log(f"  Found {len(rsp_tabs)} RSP tab(s): {rsp_tabs}")
        for tab_name in rsp_tabs:
            info = sprint_tab_info[tab_name]
            log(f"\n  '{tab_name}':")
            log(f"    Header row: {info['header_row']}")
            log(f"    Data rows: {info['data_rows']}")
            log(f"    Unique phones: {info['unique_phones']}")
            log(f"    Phones NOT in Consolidated: {len(info['phones_not_in_consolidated'])}")
            log(f"    Phone column detected: '{info['phone_col']}'")
            log(f"    Core columns: {info['core_headers']}")
            log(f"    Extra columns count: {len(info['extra_headers'])}")
            if info['nps_distribution']:
                log(f"    NPS distribution: {info['nps_distribution']}")

            # Compare with a normal sprint tab
            normal_tabs = [s for s in sprint_data_sheets
                          if s not in rsp_tabs
                          and sprint_tab_info.get(s, {}).get('core_headers')]
            if normal_tabs:
                ref_tab = normal_tabs[0]
                ref_core = sprint_tab_info[ref_tab]['core_headers']
                rsp_core = info['core_headers']
                log(f"\n    Column comparison (core) vs '{ref_tab}':")
                log(f"      Reference core cols: {ref_core}")
                log(f"      RSP core cols:       {rsp_core}")
                if ref_core == rsp_core:
                    log(f"      MATCH: Core columns are identical")
                else:
                    log(f"      DIFFERENT: Core columns differ")
    else:
        log("  No RSP tabs found.")

    # ---------- Special focus: Jan'26 tab ----------
    log_sep("=")
    log("\n>>> SPECIAL FOCUS: 'Sprint 14 Jan26' TAB <<<")
    log_sep()

    if jan26_tabs:
        for tab_name in jan26_tabs:
            info = sprint_tab_info[tab_name]
            log(f"  Sheet: '{tab_name}'")
            log(f"  Header row: {info['header_row']}")
            log(f"  Data rows: {info['data_rows']}")
            log(f"  Unique phones: {info['unique_phones']}")
            log(f"  Phones NOT in Consolidated: {len(info['phones_not_in_consolidated'])}")
            log(f"  Core columns: {info['core_headers']}")
            if info['nps_distribution']:
                log(f"  NPS distribution: {info['nps_distribution']}")
    else:
        log("  Could not find Sprint 14 Jan26 tab. Searching...")
        candidates = [s for s in sprint_data_sheets if 'jan' in s.lower()]
        if candidates:
            log(f"  Jan candidates: {candidates}")
            for c in candidates:
                info = sprint_tab_info[c]
                log(f"    '{c}': {info['data_rows']} rows, {info['unique_phones']} phones, {len(info['phones_not_in_consolidated'])} not in Consolidated")
        else:
            log("  No Jan tabs found.")

    # ---------- Summary ----------
    log_sep("=")
    log("\n>>> SUMMARY <<<")
    log_sep("=")

    log(f"\n  CONSOLIDATED TAB:")
    log(f"    Data rows: {len(consol_data)}")
    log(f"    Unique phones (10-digit normalized): {len(consol_phones)}")

    log(f"\n  ALL SPRINT DATA TABS COMBINED:")
    log(f"    Sprint data tabs analyzed: {len(sprint_tab_info)}")
    log(f"    Total data rows across sprints: {total_sprint_rows}")
    log(f"    Unique phones across all sprints: {len(all_sprint_phones)}")

    log(f"\n  GAP ANALYSIS:")
    log(f"    Phones in sprint tabs but NOT in Consolidated: {len(missing_from_consolidated)}")
    log(f"    Phones in Consolidated but NOT in any sprint tab: {len(consol_phones - all_sprint_phones)}")
    log(f"    Phones in BOTH: {len(consol_phones & all_sprint_phones)}")

    # Per-tab breakdown
    log(f"\n  PER-TAB BREAKDOWN:")
    log(f"  {'Sheet Name':<30} {'Rows':>6} {'Phones':>7} {'Missing':>8} {'In Consol':>10}")
    log(f"  {'-'*30} {'-'*6} {'-'*7} {'-'*8} {'-'*10}")

    tabs_with_missing = []
    for sheet_name, info in sprint_tab_info.items():
        missing_count = len(info['phones_not_in_consolidated'])
        in_consol = info['unique_phones'] - missing_count
        log(f"  {sheet_name:<30} {info['data_rows']:>6} {info['unique_phones']:>7} {missing_count:>8} {in_consol:>10}")
        if missing_count > 0:
            tabs_with_missing.append((sheet_name, missing_count, info['data_rows']))

    log(f"\n  TOTAL additional unique phones from sprint tabs: {len(missing_from_consolidated)}")

    # Estimate additional ROWS (not just phones) we'd gain
    # For tabs where ALL phones are missing (like RSP tabs), all rows are new
    log(f"\n  ESTIMATED ADDITIONAL ROWS:")
    total_additional_rows = 0
    for sheet_name, info in sprint_tab_info.items():
        if info['data_rows'] == 0:
            continue
        missing_pct = len(info['phones_not_in_consolidated']) / max(info['unique_phones'], 1)
        if missing_pct > 0.9:
            # Nearly all phones are missing - likely entire tab is new data
            est_rows = info['data_rows']
            log(f"    '{sheet_name}': ~{est_rows} rows (nearly all phones are new)")
            total_additional_rows += est_rows
        elif missing_pct > 0:
            # Some phones missing - estimate proportional rows
            est_rows = int(info['data_rows'] * missing_pct)
            log(f"    '{sheet_name}': ~{est_rows} rows ({missing_pct:.0%} of phones are new)")
            total_additional_rows += est_rows
    log(f"\n  TOTAL estimated additional rows: ~{total_additional_rows}")

    # Duplicate check
    log_sep("=")
    log("\n>>> DUPLICATE PHONE CHECK ACROSS SPRINT TABS <<<")
    log_sep()

    phone_to_tabs = {}
    for sheet_name, info in sprint_tab_info.items():
        for p in info['phones']:
            if p not in phone_to_tabs:
                phone_to_tabs[p] = []
            phone_to_tabs[p].append(sheet_name)

    multi_tab_phones = {p: tabs for p, tabs in phone_to_tabs.items() if len(tabs) > 1}
    log(f"  Phones appearing in MULTIPLE sprint tabs: {len(multi_tab_phones)}")
    if multi_tab_phones:
        tab_count_dist = Counter(len(tabs) for tabs in multi_tab_phones.values())
        for cnt, freq in sorted(tab_count_dist.items()):
            log(f"    Phones in exactly {cnt} tabs: {freq}")

    # Column structure comparison
    log_sep("=")
    log("\n>>> COLUMN STRUCTURE COMPARISON <<<")
    log_sep()

    # Group by core header structure
    core_structures = {}
    for name, info in sprint_tab_info.items():
        if info['core_headers']:
            key = tuple(str(h).strip().lower() for h in info['core_headers'])
            if key not in core_structures:
                core_structures[key] = {'tabs': [], 'headers': info['core_headers']}
            core_structures[key]['tabs'].append(name)

    log(f"\n  Unique CORE column structures: {len(core_structures)}")
    for i, (key, val) in enumerate(core_structures.items(), 1):
        log(f"\n  Structure {i} (used by {len(val['tabs'])} tab(s)):")
        log(f"    Tabs: {val['tabs']}")
        log(f"    Core headers: {val['headers']}")

    # Show Consolidated columns for reference
    if consol_headers:
        log(f"\n  Consolidated columns (for merge reference):")
        for i, h in enumerate(consol_headers[:30], 1):
            log(f"    {i:3d}. {h}")
        if len(consol_headers) > 30:
            log(f"    ... and {len(consol_headers) - 30} more")

    # ---------- Recommendations ----------
    log_sep("=")
    log("\n>>> RECOMMENDATIONS <<<")
    log_sep()

    if tabs_with_missing:
        log(f"\n  FINDING: {len(missing_from_consolidated)} unique phone numbers exist in sprint tabs")
        log(f"  but are MISSING from the Consolidated tab.")
        log(f"\n  TABS WITH MISSING DATA (sorted by impact):")
        for sheet_name, mc, total_rows in sorted(tabs_with_missing, key=lambda x: -x[1]):
            log(f"    '{sheet_name}': {mc} missing phones ({total_rows} total rows)")

        log(f"\n  COLUMN MAPPING for merge:")
        log(f"    Sprint core columns -> Consolidated columns:")
        log(f"      USER_ID              -> (no direct match)")
        log(f"      PROFILE_ALL_IDENTITIES -> (no direct match)")
        log(f"      USER_RATING          -> NPS")
        log(f"      NPS_CLASSIFICATION   -> NPS Group")
        log(f"      COMMENT              -> OE")
        log(f"      TIMESTAMP            -> (Sprint Start/End Date)")
        log(f"      Mobile               -> Phone Number")

        log(f"\n  RECOMMENDED NEXT STEPS:")
        log(f"    1. Build merge script to extract missing phone rows from sprint tabs")
        log(f"    2. Map sprint columns to Consolidated schema")
        log(f"    3. Derive Sprint ID from sheet name")
        log(f"    4. Append to create expanded dataset")
        log(f"    5. Re-run NPS analysis on expanded data")
    else:
        log("\n  All sprint tab phones appear to be in the Consolidated tab.")

    log("\n" + "=" * 90)
    log("END OF INVESTIGATION REPORT")
    log("=" * 90)

    wb.close()

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"\n[Report saved to: {OUTPUT_FILE}]")


if __name__ == "__main__":
    main()
